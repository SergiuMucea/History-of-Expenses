import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import os
import sys
import time
from datetime import datetime
from tqdm import tqdm

from BankProcesses.BankOperations.RowFind import find_start_row, find_last_row
from BankProcesses.BankOperations import TransactionManagement as tm
from BankProcesses.BankOperations import CellStyles as cs
from BankProcesses.THRDProcesses import FileConversion as fc
from BankProcesses.THRDProcesses import GmailHandler as gmail

data_folder = Path("path/to/the/Project/OutputFiles/")
bank_statement = data_folder / "Statements.xlsx"
final_file = data_folder / "History of Expenses.xlsx"

header_keyword = "Data\ntranzac\u0163iei"    # The ASCII form of the Romanian "Data tranzacției"


# Calculates the total of the processed statements and assigns it the category of Economii Casa
# Because all expenses done from Third Bank are basically expenses taken from this category.
def set_total_amount(ws, start_row, last_row):
    
    # Making a dictionary of items containing {month: expenses}
    months_totals = {}
    
    for i in range(start_row, last_row):
        month_cell = ws.cell(row=i, column=1)
        checked_cell = ws.cell(row=i, column=3)
        
        month = datetime.strptime(month_cell.value, "%d-%m-%Y").strftime("%m-%Y")
        
        if month not in months_totals:
            months_totals[month] = []
        
        if checked_cell.value is not None:
            months_totals[month].append(checked_cell.value)
    
    # Calculating the total of the {month: expenses} and creating an excel line with the month and amount
    for month, amounts in months_totals.items():
        total_amount = sum(amounts)
        expense_month = datetime.strptime(f"15-{month}", "%d-%m-%Y").strftime("%d-%m-%Y")
        
        ws.insert_rows(2, amount=1)
        tm.set_value(ws, 2, 1, expense_month)
        tm.set_value(ws, 2, 2, f"Total cheltuieli Third Bank: {month}\nDin contul de economii")
        if total_amount > 0:   # Fixing credit/debit amounts
            tm.set_value(ws, 2, 3, -abs(total_amount))
        else:
            tm.set_value(ws, 2, 3, abs(total_amount))
        tm.set_value(ws, 2, 4, None)
        tm.set_value(ws, 2, 5, "Economii")
        tm.set_value(ws, 2, 6, "Economii")
        tm.set_value(ws, 2, 7, "Sergiu")
        tm.set_value(ws, 2, 8, "Third Bank")


# Main function, which is imported in the GUI interface.
def main():
    # In case the procedures are not brought to an end, an error message will show in the GUI.
    global procedures_ended
    procedures_ended = False
    
    # Fetching attachments from the Third Bank statements in Gmail
    gmail.get_attachments()
    
    wb = openpyxl.Workbook()
    wb.save(bank_statement)
    wb_thrd = load_workbook(bank_statement)
    ws_thrd = wb_thrd.active
    wb_final = load_workbook(final_file)
    ws_final = wb_final['RawData']
    
    # Converting the attachments from pdf to excel
    fc.pdf_conversion(wb_thrd, ws_thrd, header_keyword)
    
    s_start_row = find_start_row(ws_thrd, header_keyword)
    s_last_row = find_last_row(ws_thrd)
    s_last_col = 4
    
    # This check makes sure that there are transactions in the processed pdf statement
    # If there are no transactions, then an error will be raised.
    # This is needed, because the final file will be formatted and everything deleted if no transactions in the statements.
    range_amount = s_last_row - s_start_row
    
    try:
        # defining file_error as global for error handling in the tkinter SecondaryGUI window
        global file_error
        file_error = False
        if range_amount < 1:
            raise FileNotFoundError
    except FileNotFoundError:
        file_error = "There are no transactions to be processed in the selected file(s)."
        print(file_error)
        os.remove(bank_statement)
        sys.exit()

    # Counts the rows in the consolidated excel statement file and inserts the same number of rows in the final file
    tm.insert_rows(ws_final, s_last_row, s_start_row)
    
    # Moving data from the statement file to the main final file and applying categories based on TransactionManagement module
    for i in tqdm(range(s_start_row, s_last_row)):
        time.sleep(0.001)
        dest_row = i
        tm.set_value(ws_final, dest_row, 7, "Sergiu")
        tm.set_bank(ws_final, dest_row, "Third Bank")
        tm.set_transactions(ws_thrd, ws_final, i, dest_row)

        for j in range (1, s_last_col):
            c = ws_thrd.cell(row=i, column=j)
            dest_col = j
            ws_final.cell(row=dest_row, column=dest_col).value = c.value
    
    # Removes redundant rows and counts them for proper range handling of set_total_amount()
    removed_rows = []
    
    # Removing redundant data and settling credit / debit amounts
    for i in tqdm(range(s_last_row, s_start_row - 1, -1)):
        time.sleep(0.001)
        dest_row = i
        remove_redundant = tm.remove_redundant(ws_thrd, ws_final, i, dest_row)
        removed_rows.extend(remove_redundant)

    ws_final_rows = s_last_row - len(removed_rows)
    set_total_amount(ws_final, s_start_row, ws_final_rows)

    # Formatting and styling the final file
    cs.format_date(ws_final)
    cs.set_accounting_format(ws_final)
    cs.wrap(ws_final)

    os.remove(bank_statement)
    wb_final.save(final_file)

    procedures_ended = True
    

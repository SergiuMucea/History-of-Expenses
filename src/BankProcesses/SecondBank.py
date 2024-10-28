from tkinter import filedialog as fd
from pathlib import Path
from openpyxl import load_workbook
from tqdm import tqdm
import time
import sys

from BankProcesses.BankOperations.RowFind import find_start_row, find_last_row
from BankProcesses.BankOperations import TransactionManagement as tm
from BankProcesses.BankOperations import CellStyles as cs

def retrieve_acc_owner(bank_statement):
    # defining file_error as global for error handling in the tkinter SecondaryGUI window
    global file_error
    file_error = False
    
    acc1 = "Sergiu"
    acc2 = "Ioana" 
    file = bank_statement
    try:
        if acc1.lower() in file.lower():
            acc = acc1
            return acc
        elif acc2.lower() in file.lower():
            acc = acc2
            return acc
        else:
            raise NameError
    except NameError:
        file_error = "Please make sure the Second Bank file is named correctly.\nIs it Ioana's or Sergiu's account?"
        print(file_error)
        sys.exit()
        
def set_acc_holder(dst_ws, dst_row, account):
    tm.set_value(dst_ws, dst_row, 7, account)


def file_selection():
    global file_error
    file_error = False
    try:
        bank_statement = fd.askopenfilename(title="Select Second Bank file")
        if bank_statement == '':
            raise FileNotFoundError
        elif not bank_statement.endswith(".xlsx"):
            raise ValueError
        return bank_statement
    except FileNotFoundError:
        file_error = "File selection canceled by user."
        print(file_error)
        sys.exit()
    except ValueError:
        file_error = "The selected file format is not supported.\nChoose an .xlsx file for file processing."
        print(file_error)
        sys.exit()


data_folder = Path("path/to/the/Project/OutputFiles/")
final_file = data_folder / "History of Expenses.xlsx"

# Main function, which is imported in the GUI interface.
def main():
    # In case the procedures are not brought to an end, an error message will show in the GUI.
    global procedures_ended
    procedures_ended = False
    
    bank_statement = file_selection()
    acc = retrieve_acc_owner(bank_statement)
    
    wb_scnd = load_workbook(bank_statement)
    ws_scnd = wb_scnd.active
    wb_final = load_workbook(final_file)
    ws_final = wb_final['RawData']

    s_start_row = find_start_row(ws_scnd, "Started Date")
    s_last_row = find_last_row(ws_scnd)
    s_last_col = 5

 
    # Settling proper arrangment of amounts in the statement file
    ws_scnd.insert_cols(7, 1)
    for i in tqdm(range(s_start_row, s_last_row)):
        time.sleep(0.001)
        checked_cell = ws_scnd.cell(row=i, column=6)
        dest_cell = ws_scnd.cell(row=i, column=7)
        if checked_cell.value > 0:
            dest_cell.value = checked_cell.value
            checked_cell.value = None
        elif checked_cell.value < 0:
            checked_cell.value = abs(checked_cell.value)
            
    # Counts the rows in the statement file and inserts the same number of rows in the final file
    tm.insert_rows(ws_final, s_last_row, s_start_row)
    
    # Deleting cols 1, 2 and 4 for correct data range
    ws_scnd.delete_cols(1, 2)
    ws_scnd.delete_cols(2)
    
    # Extending the tm.travels list according to specific items in Second Bank
    # Needed for tm.set_transactions function
    # Done separately, because otherwise interracting with Third bank
    scnd_travels = "RON"
    tm.travels.append(scnd_travels)
    
    # Moving data from the statement file to the main final file and applying categories based on TransactionManagement module
    for i in tqdm(range(s_start_row, s_last_row)):
        time.sleep(0.001)
        dest_row = i
        set_acc_holder(ws_final, dest_row, acc)
        tm.set_bank(ws_final, dest_row, "Second Bank")
        tm.set_transactions(ws_scnd, ws_final, i, dest_row)
        
        for j in range(1, s_last_col):
            c = ws_scnd.cell(row=i, column=j)
            dest_col = j
            ws_final.cell(row=dest_row, column=dest_col).value = c.value

    # Removing the specific Second Bank scnd_travels items from the original tm.travels list.
    while scnd_travels in tm.travels:
        tm.travels.remove(scnd_travels)

    # Extending the tm.excluded_trans list according to specific items in Second Bank
    excluded_scnd_trans = ["Ungureanu", "Mucea"]
    tm.excluded_trans.extend(excluded_scnd_trans)
    
    # Removing redundant data and settling credit / debit amounts
    for i in tqdm(range(s_last_row, s_start_row - 1, -1)):  # s_start_row minus 1, for end-of-range accuracy
        time.sleep(0.001)
        dest_row = i
        tm.last_procedures(ws_scnd, ws_final, i, dest_row)

    # Removing the specific Second Bank excluded_scnd_trans items from the original tm.excluded_trans list.
    for item in excluded_scnd_trans:
        while item in tm.excluded_trans:
            tm.excluded_trans.remove(item)

    # Formatting and styling the final file
    cs.format_date(ws_final)
    cs.set_accounting_format(ws_final)
    cs.wrap(ws_final)

    wb_final.save(final_file)
    
    procedures_ended = True
    

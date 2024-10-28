from tkinter import filedialog as fd
from pathlib import Path
from xls2xlsx import XLS2XLSX
from openpyxl import load_workbook
from tqdm import tqdm
import time
import sys
import os

from BankProcesses.BankOperations.RowFind import find_start_row, find_last_row
from BankProcesses.BankOperations import TransactionManagement as tm
from BankProcesses.BankOperations import CellStyles as cs

data_folder = Path("path/to/the/Project/OutputFiles/")
bank_statement = data_folder / "Statements.xlsx"
final_file = data_folder / "History of Expenses.xlsx"


# Makes the file conversion for the First Bank files which normally are received in .xls format from the bank.
def file_conv():
    # defining file_error as global for error handling in the tkinter SecondaryGUI window
    global file_error
    file_error = False
    try:
        bank_statement_old = fd.askopenfilename(title="Select First Bank file")
        if bank_statement_old == '':
            raise FileNotFoundError
        elif not bank_statement_old.endswith(".xls"):
            raise ValueError
        x2x = XLS2XLSX(bank_statement_old)
        x2x.to_xlsx(bank_statement)
        
    except FileNotFoundError:
        file_error = "File selection canceled by user."
        print(file_error)
        sys.exit()
        
    except ValueError:
        file_error = "The selected file format is not supported.\nChoose an .xls file for file conversion."
        print(file_error)
        sys.exit()
        

def set_acc_holder(src_ws, dst_ws, dst_row, dst_col=7):
    acc1 = "Sergiu"
    acc2 = "Ioana"
    if acc1 in src_ws['B1'].value:
        tm.set_value(dst_ws, dst_row, dst_col, acc1)
    elif acc2 in src_ws['B1'].value:
        tm.set_value(dst_ws, dst_row, dst_col, acc2)
    else:
        tm.set_value(dst_ws, dst_row, dst_col, "Unknown") 


# Main function, which is imported in the GUI interface.
def main():
    # In case the procedures are not brought to an end, an error message will show in the GUI.
    global procedures_ended
    procedures_ended = False
    
    file_conv()
    
    wb_frst = load_workbook(bank_statement)
    ws_frst = wb_frst.active
    wb_final = load_workbook(final_file)
    ws_final = wb_final['RawData']

    s_start_row = find_start_row(ws_frst, "Data")
    s_last_row = find_last_row(ws_frst, 2)  # There are 3 redundant rows at the end of the First Bank statement
    s_last_col = 4

    # Counts the rows in the statement file and inserts the same number of rows in the final file
    tm.insert_rows(ws_final, s_last_row, s_start_row)

    # Moving data from the statement file to the main final file and applying categories based on TransactionManagement module
    for i in tqdm(range(s_start_row, s_last_row)):
        time.sleep(0.001)
        dest_row = i - s_start_row + 2    # Plus 2, because otherwise, dest_row would be 0
        set_acc_holder(ws_frst, ws_final, dest_row)
        tm.set_bank(ws_final, dest_row, "First Bank")
        tm.set_transactions(ws_frst, ws_final, i, dest_row)

        for j in range (1, s_last_col + 1):
            selected_cell = ws_frst.cell(row=i, column=j)
            dest_col = j
            ws_final.cell(row=dest_row, column=dest_col).value = selected_cell.value

    # Removing redundant data and settling credit / debit amounts
    for i in tqdm(range(s_last_row, s_start_row - 1, -1)):   # s_start_row minus 1, for end-of-range accuracy
        time.sleep(0.001)
        dest_row = i - s_start_row + 2
        tm.last_procedures(ws_frst, ws_final, i, dest_row)

    # Formatting and styling the final file
    cs.format_date(ws_final)
    cs.set_accounting_format(ws_final)
    cs.wrap(ws_final)
    
    os.remove(bank_statement)
    wb_final.save(final_file)
    
    procedures_ended = True

